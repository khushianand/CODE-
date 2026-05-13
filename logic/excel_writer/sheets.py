"""Excel sheet-level writers for data, dashboard, and disposition sheets."""

from __future__ import annotations
from typing import Optional
import pandas as pd
#from openpyxl.chart import PieChart, Reference
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

#from logic.excel_writer.formatting import BOLD, BLUE, CENTER, WHITE, auto_width, write_headers, write_summary_headers
from logic.excel_writer.formatting import (
    BLACK_THIN_BORDER,
    BOLD,
    BLUE,
    CENTER,
    DATA_ALIGNMENT,
    WHITE,
    apply_table_formatting,
    auto_width,
    write_headers,
    write_summary_headers,
)
from logic.highlight_logic import severity_fill
from logic.parser import TEMPLATE_COLUMNS
#from logic.summary_generator import severity_summary
from logic.summary_generator import disposition_summary, expert_severity_summary, severity_chart_summary, severity_summary
from logic.excel_writer.three_uk_qualys import (
    THREE_UK_QUALYS_TOTAL_COLUMNS,
)

DASHBOARD_SOURCE_START_COL = 27  # AA: off-screen chart source data, keeping visible dashboard chart-only.
SEVERITY_CHART_COLORS = ["9B0F06", "D53E0F", "F77F00", "FCBF49"]
BAR_CHART_COLORS = ["4472C4", "C00000", "F4B183", "C00000", "70AD47", "7030A0", "8064A2", "92D050"]
WRAP_CENTER = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)
DISPOSITION_ORDER = [
    "Transferred",
    "Mitigate in Future Release",
    "Mitigate",
    "In Analysis",
    "False Positive/Not Relevant",
    "Accept/No Solution Expected",
    "Accept",
    "False Positive",
]



def _write_data_rows(ws, df: pd.DataFrame):
    for row_idx, row in enumerate(df[TEMPLATE_COLUMNS].itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        risk_cell = ws.cell(row=row_idx, column=4)
        risk_cell.fill = severity_fill(risk_cell.value)
        risk_cell.font = BOLD


def write_main_sheet(
    ws,
    df: pd.DataFrame,
    project: str,
    scanner: str,
):

    write_headers(ws)

    ws["B1"] = project
    ws["E1"] = scanner

    _write_data_rows(ws, df)

    ws.freeze_panes = "A3"

    auto_width(ws)#def write_main_sheet(ws, df: pd.DataFrame, project: str, scanner: str):


    # Use the Qualys-style Total Data layout only for the exact requested
    # combination: project 3UK + scanner Qualys + Total Data sheet. Every other
    # project/scanner/sheet combination falls through to the universal template below.

    

def _clear_dashboard(ws):
    ws._charts = []
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    """for offset, row in enumerate(df.itertuples(index=False), start=0):
        r = start_row + 2 + offset

        key_cell = ws.cell(r, start_col, row[0])
        ws.cell(r, start_col + 1, row[1])

        # ✅ Apply your existing severity color logic
        if "severity" in str(df.columns[0]).lower():
            key_cell.fill = severity_fill(key_cell.value)
            key_cell.font = BOLD
    """

def _hide_chart_source_columns(ws, start_col: int, width: int):
    for col_idx in range(start_col, start_col + width):
        col = ws.column_dimensions[get_column_letter(col_idx)]
        col.hidden = False
        col.width = 1  
    """Keep source data off-screen but plottable in Excel/LibreOffice.
    Fully hidden columns can make some spreadsheet viewers render blank charts even
    when chart.plotVisOnly is False, so keep the columns technically visible with
    very small width and white text.
    """                  

def _write_chart_source(ws, df: pd.DataFrame, start_row: int, start_col: int) -> int:
    for col_offset, col_name in enumerate(df.columns):
        ws.cell(start_row, start_col + col_offset, col_name)
    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        for col_offset, value in enumerate(row):
            ws.cell(start_row + row_offset, start_col + col_offset, value)
    return len(df)



def _color_series_points(series, colors):
    series.data_points = []
    for idx, color in enumerate(colors):
        point = DataPoint(idx=idx)
        point.graphicalProperties.solidFill = color
        series.data_points.append(point)


def _add_pie(ws, title: str, source_row: int, source_col: int, size: int, anchor: str):    
    chart = PieChart()
    chart.title = title
#    data = Reference(ws, min_col=table_col + 1, min_row=table_row + 1, max_row=table_row + 1 + size)
#    cats = Reference(ws, min_col=table_col, min_row=table_row + 2, max_row=table_row + 1 + size)
    data = Reference(ws, min_col=source_col + 1, min_row=source_row, max_row=source_row + size)
    cats = Reference(ws, min_col=source_col, min_row=source_row + 1, max_row=source_row + size)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        _color_series_points(chart.series[0], SEVERITY_CHART_COLORS[:size])
    chart.height = 7
    chart.width = 12
    chart.plotVisOnly = False
    ws.add_chart(chart, anchor)




def _add_bar(ws, title: str, source_row: int, source_col: int, rows: int, cols: int, anchor: str):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Count"
    chart.x_axis.title = ""
    data = Reference(ws, min_col=source_col + 1, max_col=source_col + cols - 1, min_row=source_row, max_row=source_row + rows)
    cats = Reference(ws, min_col=source_col, min_row=source_row + 1, max_row=source_row + rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for idx, series in enumerate(chart.series):
        color = BAR_CHART_COLORS[idx % len(BAR_CHART_COLORS)]
        series.graphicalProperties.solidFill = color
    if len(chart.series) == 1 and rows > 1:
        _color_series_points(chart.series[0], BAR_CHART_COLORS[:rows])
    chart.height = 7
    chart.width = 14
    chart.plotVisOnly = False
    ws.add_chart(chart, anchor)

def _append_pie_charts(ws, total_df: pd.DataFrame, unique_df: pd.DataFrame, source_col: int = DASHBOARD_SOURCE_START_COL) -> int:
    total_table = severity_chart_summary(total_df, include_total=False)
    unique_table = severity_chart_summary(unique_df, include_total=False)

    total_row = 3
    unique_row = total_row + len(total_table) + 3
    total_size = _write_chart_source(ws, total_table, total_row, source_col)
    unique_size = _write_chart_source(ws, unique_table, unique_row, source_col)

    _add_pie(ws, "Total Vulnerabilities per severity", total_row, source_col, total_size, "A3")
    _add_pie(ws, "Unique Vulnerabilities per severity", unique_row, source_col, unique_size, "A20")
    return unique_row + unique_size + 3


def _append_vams_bar_charts(ws, vams_df: pd.DataFrame, source_row: int, source_col: int = DASHBOARD_SOURCE_START_COL):
    reported_expert = expert_severity_summary(vams_df)
    disposition = disposition_summary(vams_df, DISPOSITION_ORDER)

    reported_row = source_row
    disposition_row = reported_row + len(reported_expert) + 3
    reported_rows = _write_chart_source(ws, reported_expert, reported_row, source_col)
    disposition_rows = _write_chart_source(ws, disposition, disposition_row, source_col)

    _add_bar(
        ws,
        "Reported Severity Vs Expert Severity",
        reported_row,
        source_col,
        reported_rows,
        len(reported_expert.columns),
        "M3",
    )
    _add_bar(
        ws,
        "Disposition",
        disposition_row,
        source_col,
        disposition_rows,
        len(disposition.columns),
        "M20",
    )


def write_summary_sheet(
    ws,
    new_df,
    old_df,
    unique_df,
    project,
    scanner,
    include_old_summary: bool = True,
    include_vams_charts: bool = False,
    vams_chart_df: pd.DataFrame | None = None,
):
    """Write the chart-only Dashboard sheet.

    Tab 1 and Tab 2 always show total/unique severity pie charts. VAMS-only
    bar charts are appended when Tab 3 refreshes the dashboard after enrichment.
    Hidden AA+ source data powers the charts without displaying summary tables.
    """
    _clear_dashboard(ws)
    write_summary_headers(ws, project, scanner)
    next_source_row = _append_pie_charts(ws, new_df, unique_df)
    if include_vams_charts:
        _append_vams_bar_charts(ws, vams_chart_df if vams_chart_df is not None else new_df, next_source_row)
    _hide_chart_source_columns(ws, DASHBOARD_SOURCE_START_COL, 5)


    auto_width(ws)


def write_disposition_sheet(ws, project: str = "", scanner: str = "", *_, **__):
    """Create static Disposition sheet template included in every output workbook."""

    # Row 1 project/scanner values
    ws["A1"] = "Project Name:"
    ws["D1"] = "Scanner:"
    ws["B1"] = project
    ws["E1"] = scanner

    # Apply row-1 color coding/styling across the entire first row.
    for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = BLUE
        cell.font = WHITE
        cell.alignment = CENTER
        cell.border = BLACK_THIN_BORDER

    headers = ["Disposition Value", "Description"]
    for col_idx, title in enumerate(headers, start=1):
        # Table starts at row 3 (row 2 intentionally left blank)
        cell = ws.cell(row=3, column=col_idx, value=title)
        cell.fill = BLUE
        cell.font = WHITE
        cell.alignment = CENTER
        cell.border = BLACK_THIN_BORDER

    # 👉 EDIT HERE (COLUMN A VALUES):
    # Each tuple is: ("Disposition Value", "Description")
    # - First value goes to column A (Disposition Value).
    # - Second value goes to column B (Description).
    # Example:
    # ("Mitigate", "Fix planned in next quarterly release.")
    disposition_rows = [
        (
            "Transferred",
            "Responsibility for resolution is assumed by Customer (mainly used for scans). This is used for configuration or certificate related issues which typically need to be addressed by the customer.",
        ),
        (
            "Mitigate in Future Release",
            "Vulnerability will be mitigated but not in current product/release where vulnerability was reported (i.e. if not mitigated in this product/release, it will be fixed in a later release).",
        ),
        ("Mitigate", "Vulnerability will be fixed in current product/release configured in VAMS."),
        ("In Analysis", "Assessment in progress - intermediate state before final disposition value."),
        (
            "False Positive/Not Relevant",
            "False Positive could represent a known scanner issue or Not Relevant notification against 3rd party component used in product. False Positive should be used only when there is zero risk that the vulnerability exists.",
        ),
        (
            "Accept/No Solution Expected",
            "Risk is accepted. Used for situations where vulnerability risk is low and impact to release may be high, or when vendor states they will not provide a patch/fix.",
        ),
    ]

    for row_idx, (value, description) in enumerate(disposition_rows, start=4):
        ws.cell(row=row_idx, column=1, value=value)
        ws.cell(row=row_idx, column=2, value=description)

    # 👉 OPTIONAL: adjust row heights if you want larger cells for long descriptions.
    # Example: ws.row_dimensions[3].height = 70
    #ws.row_dimensions[4].height = 60
    #ws.row_dimensions[6].height = 45
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BLACK_THIN_BORDER
            if cell.row in [1, 3]:
                cell.alignment = WRAP_CENTER
            else:
                cell.alignment = DATA_ALIGNMENT

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 15

    auto_width(ws)
